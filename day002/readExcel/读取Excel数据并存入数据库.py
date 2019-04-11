# _*_ coding:UTF-8 _*_
# 开发人员: 理想国真恵玩-张大鹏
# 开发团队: 理想国真恵玩
# 开发时间: 2019-04-11 10:14
# 文件名称: 读取Excel数据并存入数据库.py
# 开发工具: PyCharm

import os  # 系统操作库
import sqlite3  # 简易数据库
from openpyxl import load_workbook  # 读取Excel的工具

# 商人?
merchants = {}  # 定义一个空字典
if not os.path.isfile('merchant.sqlite'):
    conn = sqlite3.connect("merchant.sqlite")  # 连接数据库实例
    c = conn.cursor()  # 操作数据库的游标
    # 创建一张表 叫merchant
    c.execute(
        """
        create table if not exists merchant(
        id integer primary key autoincrement,
        name text not null,
        minimum integer not null 
        )
        """
    )
    # 创建另一张表 叫transactions
    c.execute(
        """
        create table if not exists `transactions`(
        id integer primary key autoincrement,
        merchant text not null ,
        total double not null ,
        need_to_pay integer not null 
        )
        """
    )
    # 执行创建
    conn.commit()
else:
    conn = sqlite3.connect("merchant.sqlite")
    c = conn.cursor()
    # 查询merchant表中的所有数据
    c.execute("select * from merchant")
    for row in c.fetchall():
        print('--' * 22)
        print(row)
        print('--' * 22)
        merchants[row[1]] = row[2]
# 定义用户选择
choice = ''


def add_merchant():
    """
    添加数据
    :return:
    """
    merchant = input("\n请输入商人名字:")
    # 这里的eval的作用是将用户输入的数据(本身是字符串)转换为整数或浮点数
    minimum = eval(input("\n请输入最小数量:"))
    print('--' * 22)
    print("minimum被转换后:", minimum)  # 打印查看minimum被转换成了什么
    print('--' * 22)
    # 连接数据库,并将新的数据添加到数据库
    conn = sqlite3.connect("merchant.sqlite")
    c = conn.cursor()
    c.execute("insert into `merchant` (name,minimum) values('{}',{})".format(
        merchant, minimum
    ))
    conn.commit()
    # merchants的基本数据结构是 {'商人名字':最小数量}
    merchants[merchant] = minimum
    print("数据添加成功.")


def read_excel():
    """
    读取Excel表
    :return:
    """
    filepath = input("\n请输入xlsx文件的路径(相对路径或绝对路径都可以):")
    wb = load_workbook(filename=filepath)  # 根据路径读取
    active = wb.active  # 类似于游标
    transcations = {}  # 另一张表对应的字典
    for row in active.rows:
        print('--' * 22)
        print("测试一下读取的Excel表每一行是什么:", row)
        print('--' * 22)
        if row[0].value == "Number":
            continue  # 跳过
        if row[2].value not in transcations:
            # 空字典 这个是为了避免重复
            transcations[row[2].value] = {"amount": float(row[4].value)}
        else:
            # 新增一行
            transcations[row[2].value]['amount'] += float(row[4].value)
    for item in transcations:
        print('--' * 22)
        print("测试transcations中装的是什么:", item)
        print('--' * 22)
        # 遍历字典
        if item not in merchants:
            # 不在商人表中
            transcations[item]["need_to_pay"] = 0
            print("%s 是错误的商人" % item)
        else:
            merchant = merchants[item]
            if transcations[item]['amount'] >= merchant:
                transcations[item]["need_to_pay"] = 0
                print("%s:\n最小数量:%.2f\n总额:%.2f\n" % (
                    item, merchant, transcations[item]['amount']
                ))
            else:
                transcations[item]["need_to_pay"] = 1
                print("%s:\n最小数量:%.2f\n总共:%.2f需要支付\n" % (
                    item, merchant, transcations[item]['amount']
                ))
        conn = sqlite3.connect("merchant.sqlite")
        c = conn.cursor()
        c.execute("insert into `transactions` (merchant,total,need_to_pay) values('%s',%i,%i)" % (
            item, transcations[item]['amount'], transcations[item]['need_to_pay']
        ))
        conn.commit()


def read_excel_and_show():
    filepath = input("\n请输入xlsx文件路径:  ")
    wb = load_workbook(filename=filepath)
    active = wb.active

    transactions = {}
    for row in active.rows:
        if row[0].value == "Number":
            continue

        if row[2].value not in transactions:
            transactions[row[2].value] = {"amount": float(row[4].value)}
        else:
            transactions[row[2].value]["amount"] += float(row[4].value)

    for item in transactions:
        if item not in merchants:
            transactions[item]["need_to_pay"] = 0

            print(("%s 是错误的商人" % item))
        else:
            merchant = merchants[item]

            if transactions[item]["amount"] >= merchant:
                transactions[item]["need_to_pay"] = 0
                print(("%s:\n最小数量:%.2f\n总额: %.2f\n" %
                       (item, merchant, transactions[item]["amount"])))
            else:
                transactions[item]["need_to_pay"] = 1
                print(("%s:\n最小数量:%.2f\n已支付: %.2f\n %.2f 需要支付\n" %
                       (item, merchant, transactions[item]["amount"]), merchant - transactions[item]["amount"]))

        conn = sqlite3.connect("merchant.sqlite")
        c = conn.cursor()

        c.execute("INSERT INTO `transactions` (merchant, total, need_to_pay) VALUES('%s', %i, %i)" %
                  (item, transactions[item]["amount"], transactions[item]["need_to_pay"]))
        conn.commit()


while choice != 'q':
    print("\n-------------------------------------------")
    print("-         读取Excel表                   -")
    print("-------------------------------------------")

    print("[1] 读取 Excel")
    print("[2] 添加商人")
    print("[3] 读取Excel文件并展示需要支付的商人")
    print("[q] 退出")

    choice = input(
        "\n请输入操作符(1/2/3/q):  ")

    if choice == '1':
        read_excel()
    elif choice == '2':
        add_merchant()
    elif choice == '3':
        read_excel_and_show()
    elif choice == 'q':
        print("您已退出系统!")
    else:
        read_excel()
