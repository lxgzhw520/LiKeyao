import os
import sqlite3
from openpyxl import load_workbook

merchants = {}
if not os.path.isfile("merchant.sqlite"):
    conn = sqlite3.connect("merchant.sqlite")
    c = conn.cursor()

    c.execute(
        '''CREATE TABLE IF NOT EXISTS merchant(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            minimum INTEGER NOT NULL
        )''')
    c.execute(
        '''CREATE TABLE IF NOT EXISTS `transactions`(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            merchant TEXT NOT NULL,
            total DOUBLE NOT NULL,
            need_to_pay INTEGER NOT NULL
        )''')
    conn.commit()
else:
    conn = sqlite3.connect("merchant.sqlite")
    c = conn.cursor()

    c.execute("SELECT * from merchant")

    for row in c.fetchall():
        merchants[row[1]] = row[2]

choice = ''


def add_merchant():
    merchant = input("\nPlease enter merchant name:  ")
    minimum = eval(input("\nPlease enter minimum amount:  "))

    conn = sqlite3.connect("merchant.sqlite")
    c = conn.cursor()

    c.execute("INSERT INTO `merchant` (name, minimum) VALUES('%s', %i)" %
              (merchant, minimum))
    conn.commit()

    merchants[merchant] = minimum

    print(("Add %s to database." % merchant))


def read_excel():
    filepath = input("\nPlease enter xlsx filepath:  ")
    wb = load_workbook(filename=filepath)
    active = wb.active

    transactions = {}
    for row in active.rows:
        if row[0].value == "Number":
            continue

        if row[2].value not in transactions:
            transactions[row[2].value] = {"amount": float(row[4].value)}
        else:
            transactions[row[2].value]["amount"] += float(row[4].value)

    for item in transactions:
        if item not in merchants:
            transactions[item]["need_to_pay"] = 0

            print(("%s is a wrong merchant" % item))
        else:
            merchant = merchants[item]

            if transactions[item]["amount"] >= merchant:
                transactions[item]["need_to_pay"] = 0
                print(("%s:\nminimum:%.2f\ntotal amount: %.2f\n" %
                       (item, merchant, transactions[item]["amount"])))
            else:
                transactions[item]["need_to_pay"] = 1
                print(("%s:\nminimum:%.2f\ntotal amount: %.2f need to pay\n" %
                       (item, merchant, transactions[item]["amount"])))

        conn = sqlite3.connect("merchant.sqlite")
        c = conn.cursor()

        c.execute("INSERT INTO `transactions` (merchant, total, need_to_pay) VALUES('%s', %i, %i)" %
                  (item, transactions[item]["amount"], transactions[item]["need_to_pay"]))
        conn.commit()


def read_excel_and_show():
    filepath = input("\nPlease enter xlsx filepath:  ")
    wb = load_workbook(filename=filepath)
    active = wb.active

    transactions = {}
    for row in active.rows:
        if row[0].value == "Number":
            continue

        if row[2].value not in transactions:
            transactions[row[2].value] = {"amount": float(row[4].value)}
        else:
            transactions[row[2].value]["amount"] += float(row[4].value)

    for item in transactions:
        if item not in merchants:
            transactions[item]["need_to_pay"] = 0

            print(("%s is a wrong merchant" % item))
        else:
            merchant = merchants[item]

            if transactions[item]["amount"] >= merchant:
                transactions[item]["need_to_pay"] = 0
                print(("%s:\nminimum:%.2f\ntotal amount: %.2f\n" %
                       (item, merchant, transactions[item]["amount"])))
            else:
                transactions[item]["need_to_pay"] = 1
                print(("%s:\nminimum:%.2f\npayed: %.2f\n %.2f need to pay\n" %
                       (item, merchant, transactions[item]["amount"]), merchant - transactions[item]["amount"]))

        conn = sqlite3.connect("merchant.sqlite")
        c = conn.cursor()

        c.execute("INSERT INTO `transactions` (merchant, total, need_to_pay) VALUES('%s', %i, %i)" %
                  (item, transactions[item]["amount"], transactions[item]["need_to_pay"]))
        conn.commit()


while choice != 'q':
    print("\n-------------------------------------------")
    print("-          Excel Reader                   -")
    print("-------------------------------------------")

    print("[1] Read Excel")
    print("[2] Add Merchant")
    print("[3] Read Excel and show need to pay merchant")
    print("[q] Quit")

    choice = input(
        "\nPlease select an action [default action is Read Excel]:  ")

    if choice == '1':
        read_excel()
    elif choice == '2':
        add_merchant()
    elif choice == '3':
        read_excel_and_show()
    elif choice == 'q':
        print("Byebye")
    else:
        read_excel()
